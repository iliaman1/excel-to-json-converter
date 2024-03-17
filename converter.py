import datetime
import json
import logging
from dataclasses import dataclass

import openpyxl

logging.basicConfig(level=logging.INFO, format='[%(asctime)s] %(message)s')

DATA_START_ROW = 5
DATA_OFFSET = 9
MONTH_START_COLUMN = 2
MONTH_END_COLUMN = 13
TAXES_COUNT = 8
HEAD = (1, 2, 6, 8, 10)

UNP = 700069297
RECORDS_PER_PACK = 200


def batch(iterable, n=1):
    l = len(iterable)
    for ndx in range(0, l, n):
        yield iterable[ndx:min(ndx + n, l)]


@dataclass
class Month:
    d_201: float
    tax: float
    b_610: float
    b_600: float
    B_620: float
    c_650: float
    o_660: float
    mat_pom: float


@dataclass
class RawData:
    number: int
    full_name: str
    passport_number: str
    personal_number: str
    address: str
    months: list[Month]

    def get_full_name(self) -> tuple[str, str, str]:
        values = self.full_name.strip().split(' ')

        if len(values) != 3:
            logging.warning(
                f'Для "{self.full_name.strip()}" указано не полное ФИО. '
                'Будут использованы пустые значения.'
            )

        match len(values):
            case 1:
                return values[0], '', ''
            case 2:
                return values[0], values[1], ''
            case 3:
                return values[0], values[1], values[2]
            case _:
                return '', '', ''

    def generate_tar4(self) -> list[dict]:
        return [
            {
                'nmonth': index + 1,
                'nsummonth': self.months[index].d_201,
                'tar4sum': [{'ncode': 201, 'nsum': self.months[index].d_201}]
            }
            for index in range(12)
        ]

    def generate_tar7(self) -> list[dict]:
        return [
            {
                'nmonth': index + 1,
                'nsummonth': self.months[index].b_600,
                'tar7sum': [{'ncode': 600, 'nsum': self.months[index].b_600}]
            }
            for index in range(12)
        ]

    def generate_tar14(self) -> list[dict]:
        return [
            {
                'nmonth': index + 1,
                'nsumdiv': 0,
                'nsumt': self.months[index].tax
            }
            for index in range(12)
        ]

    def to_dict(self):
        last_name, first_name, surname = self.get_full_name()

        return {
            'docagentinfo': {
                'cln': self.personal_number,
                'cstranf': '112',  # 2.5 Код страны гражданства (подданства)
                'cvdoc': '01',  # 2.6 Код документа, удостоверяющего личность
                'nrate': 13,  # 3. Размер ставки подоходного налога с физических лиц, проценты
                'vfam': last_name,
                'vname': first_name,
                'votch': surname
            },
            'nsumstand': round(sum([nalog.b_600 for nalog in self.months]), 2),
            'ntsumbank': 0,
            'ntsumcalcincome': round(sum([nalog.tax for nalog in self.months]), 2),
            'ntsumcalcincomediv': 0,
            'ntsumexemp': 0,
            'ntsumincome': round(sum([nalog.d_201 for nalog in self.months]), 2),
            'ntsumnotcalc': 0,
            'ntsumprof': 0,
            'ntsumprop': 0,
            'ntsumsec': 0,
            'ntsumsoc': 0,
            'ntsumtrust': 0,
            'ntsumwithincome': 0,
            'ntsumwithincomediv': 0,
            'tar14': self.generate_tar14(),
            'tar4': self.generate_tar4(),
            'tar7': self.generate_tar7()
        }


class Converter:

    @staticmethod
    def create_raw_data_list(excel_filename: str) -> list[RawData]:
        data: list[RawData] = []
        wb = openpyxl.load_workbook(excel_filename)
        sheet = wb.active
        max_rows = sheet.max_row

        for xlsx_row in range(DATA_START_ROW, max_rows + 1, DATA_OFFSET):
            months = []
            for month in range(MONTH_START_COLUMN, MONTH_END_COLUMN + 1):
                months.append(
                    Month(
                        *[sheet.cell(row=xlsx_row + index + 1, column=month).value or 0 for index in range(TAXES_COUNT)]
                    )
                )

            data.append(
                RawData(
                    *[sheet.cell(row=xlsx_row, column=i).value for i in HEAD],
                    months=months
                )
            )

        return data

    @staticmethod
    def to_dict(data: list[RawData]) -> dict:
        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime("%Y-%m-%dT%H:%M:%S")

        return {
            'pckagent': {
                'docagent': [person.to_dict() for person in data],
                'pckagentinfo': {
                    'dcreate': formatted_time,
                    'ngod': current_time.year,
                    'nmns': 741,
                    'nmnsf': 741,
                    'ntype': 1,
                    'vexec': 'Буйко Т.С.',
                    'vphn': '72-30-50',
                    'vunp': '700069297'
                }
            }
        }

    @staticmethod
    def generate_filename(unp, form_type, department_code, part_number=None) -> str:
        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime("%Y%m%d%H%M%S")
        filename = f"D{unp}_{current_time.year}_{form_type}_{department_code}_{formatted_time}"

        if part_number is not None:
            filename += f"_{part_number:04d}"

        filename += ".json"
        return filename

    def make_files(self, data: list[RawData]):
        pages_count = len(data) // RECORDS_PER_PACK

        if pages_count > 0:
            for part, group in enumerate(batch(data, RECORDS_PER_PACK)):
                part_data = self.to_dict(group)
                filename = f'gen_json/{self.generate_filename(UNP, 1, 0, part + 1)}'

                with open(filename, 'w', encoding='utf-8') as file:
                    json.dump(part_data, file, indent=4, ensure_ascii=False)
        else:
            data = self.to_dict(data)
            filename = f'gen_json/{self.generate_filename(UNP, 1, 0)}'

            with open(filename, 'w', encoding='utf-8') as file:
                json.dump(data, file, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    convert_rti = Converter()
    rawdata = Converter.create_raw_data_list('доход2023.xlsx')
    convert_rti.make_files(rawdata)
